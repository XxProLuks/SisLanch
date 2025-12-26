"""
LANCH - Export Service
"""

import io
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def export_to_excel(data: List[Dict[str, Any]], competencia) -> io.BytesIO:
    """
    Export consumption data to Excel with formatting
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Competência {competencia.mes:02d}-{competencia.ano}"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = f"Relatório de Desconto em Folha - {competencia.mes:02d}/{competencia.ano}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Headers
    headers = ["Matrícula", "Nome", "Setor", "Centro de Custo", "Valor Total", "Competência"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data
    total_geral = 0
    for row_idx, item in enumerate(data, 4):
        ws.cell(row=row_idx, column=1, value=item["matricula"]).border = thin_border
        ws.cell(row=row_idx, column=2, value=item["nome"]).border = thin_border
        ws.cell(row=row_idx, column=3, value=item["setor"]).border = thin_border
        ws.cell(row=row_idx, column=4, value=item["centro_custo"]).border = thin_border
        
        valor_cell = ws.cell(row=row_idx, column=5, value=item["valor_total"])
        valor_cell.number_format = 'R$ #,##0.00'
        valor_cell.border = thin_border
        
        ws.cell(row=row_idx, column=6, value=item["competencia"]).border = thin_border
        total_geral += item["valor_total"]
    
    # Total row
    total_row = len(data) + 4
    ws.cell(row=total_row, column=4, value="TOTAL GERAL:").font = Font(bold=True)
    total_cell = ws.cell(row=total_row, column=5, value=total_geral)
    total_cell.font = Font(bold=True)
    total_cell.number_format = 'R$ #,##0.00'
    
    # Column widths
    column_widths = [15, 40, 25, 20, 15, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer


def export_to_csv(data: List[Dict[str, Any]], competencia) -> str:
    """
    Export to CSV compatible with TOTVS RM
    Layout: MATRICULA;NOME;CENTRO_CUSTO;VALOR;COMPETENCIA;TIPO_DESCONTO
    """
    lines = []
    
    # Header
    lines.append("MATRICULA;NOME;CENTRO_CUSTO;VALOR;COMPETENCIA;TIPO_DESCONTO")
    
    # Data
    for item in data:
        valor_str = f"{item['valor_total']:.2f}".replace(".", ",")
        comp_str = f"{competencia.mes:02d}/{competencia.ano}"
        
        line = ";".join([
            item["matricula"],
            item["nome"],
            item["centro_custo"],
            valor_str,
            comp_str,
            "LANCHONETE"  # Tipo de desconto fixo
        ])
        lines.append(line)
    
    return "\n".join(lines)
